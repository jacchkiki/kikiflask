from openpyxl import Workbook
import os.path

class exlibs:
    excelname="" 
    def __init__(self,name):  
        self.excelname=name
        self.wb = Workbook() 
       # 
        if(os.path.exists(name)): 
            self.wb.load_workbook(filename = name)
        else: 
            self.create()
 
    def create(self): 
        #ws=self.create_sheet(title="Data")
        #ws["A1"]=123
        self.wb.save(self.excelname)
    
    def test():
        v=self.wb["sheet1"]
        print(v["A1"])
    #def write(self,cell,row,value):
        
    